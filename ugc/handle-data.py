import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import os
import json
from openpyxl.utils import get_column_letter

file_index = 1
dataset = ''

with open(f'{file_index}.json', 'r', encoding='utf-8') as file:
    content = file.read()
    dataset = json.loads(content)

header_name_map = {
      'scene_id': '场景ID',
      'img_url': 'AI人物图片',
      'script_name': '剧本名',
      'story_desc': '剧情描述',
      'ai_character_desc': 'AI人物描述',
      'user_character_desc': '用户人物描述',
      'ai_open_speech': 'AI开场白',
      'scene_status': '场景状态',
      'operate': '操作',
      'reason': '不推荐理由',
      'remark': '备注'
}

operate_map = {
    '1': '推荐',
    '2': '不推荐',
    '3': '下架'
}

choice_keys = ['operate']
text_keys = ['reason', 'remark']
output_dataset = []
output_header = []

wb = openpyxl.Workbook()
ws = wb.active
ws.append(['场景ID','AI人物图片','剧本名','剧情描述','AI人物描述','用户人物描述','AI开场白','场景状态','操作','不推荐理由','备注'])

for index, data in enumerate(dataset):
    item = {}
    annotations_result = data['annotations'][0]['result']
    source_data = data['data']
    for key in source_data.keys():
        item[header_name_map[key]] = source_data[key]
    for result in annotations_result:
        key_name = result['from_name']
        if key_name in choice_keys:
            value = result['value']['choices'][0]
            item[header_name_map[key_name]] = operate_map[value]
        elif key_name in text_keys:
            value = result['value']['text'][0]
            item[header_name_map[key_name]] = value
    row_data = []
    for key in item.keys():
        if key == 'AI人物图片':
            row_data.append('')
        else:
            row_data.append(item[key])
    ws.append(row_data)
    img_path = os.path.join(f'{os.getcwd()}/output_images/3/{index+1}.png')
    img = Image(img_path)
    img.width,img.height = 80,80
    row = ws.row_dimensions[index+2]
    row.height = 50
    ws.add_image(img, f'B{index+2}')

for i in range(1,11):
    col = ws.column_dimensions[get_column_letter(i)]
    col.width = 20

wb.save(r'output.xlsx')
            